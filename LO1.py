import tkinter as tk
from tkinter import messagebox
import openpyxl
import os
import win32com.client  # To access Outlook
from bs4 import BeautifulSoup  # For parsing HTML emails

import tkinter as tk
from tkinter import* 
from tkinter import ttk

import tkinter as tk

from tkinter.ttk import Combobox
from tkinter import messagebox

from tkcalendar import Calendar, DateEntry


# Global Excel file paths
extracted_email_file = 'New_job_openings_data.xlsx'
updated_data_file = 'updated_data.xlsx'

# Function to extract and format HTML content from email body
def extract_template_data(sheet, row_num, body, sender, received_time, sender_name):
    # Parse the HTML content
    soup = BeautifulSoup(body, 'html.parser')

    # Find the table with the job details (adjust based on your actual email structure)
    table = soup.find('table')

    if table:
        # Find all rows in the table (each row has two columns: Heading and Value)
        data_dict = {}
        rows = table.find_all('tr')
        for row in rows:
            cols = row.find_all('td')
            if len(cols) == 2:
                heading = cols[0].get_text(strip=True)
                value = cols[1].get_text(strip=True)
                data_dict[heading] = value

        # Write the extracted data into Excel (ensure order is consistent)
        sheet.cell(row=row_num, column=1, value=sender)  # Sender email in column 1
        sheet.cell(row=row_num, column=2, value=received_time)
        sheet.cell(row=row_num, column=3, value=data_dict.get('CSR Name', ''))
        sheet.cell(row=row_num, column=4, value=data_dict.get('Number', ''))
        sheet.cell(row=row_num, column=5, value=data_dict.get('Action Required', ''))
        sheet.cell(row=row_num, column=6, value=data_dict.get('Brief Description', ''))
        sheet.cell(row=row_num, column=7, value=sender_name)   

    return row_num

# Helper function to process all unread emails
def process_unread_emails(unread_messages, sheet):
    row_num = sheet.max_row + 1  # Start appending at the next available row

    for message in unread_messages:
        try:
            # Extract email sender and HTML body
            sender = message.SenderEmailAddress
            sender_name = message.SenderName
            body = message.HTMLBody  # Assuming the body is in HTML format
            received_time = message.ReceivedTime.strftime('%Y-%m-%d %H:%M:%S')

            # Extract and save the data from the email
            row_num = extract_template_data(sheet, row_num, body, sender, received_time, sender_name)

            # Mark the email as read after processing
            message.UnRead = False

        except Exception as e:
            print(f"Error processing email from {sender}: {e}")

# Get all unread emails and save them into the Excel file
def get_emails():
    # Load or create the Excel file for email data
    if os.path.exists(extracted_email_file):
        wb = openpyxl.load_workbook(extracted_email_file)
        sheet = wb.active
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        # If creating new, add the header
        headers = ['Sender', 'Received Time', "CSR Name", "Number", "Action Required", "Brief Description", "Sender Name" 
                   ]
        sheet.append(headers)

    try:
        # Create a connection to Outlook
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        inbox = outlook.Folders("preetamjangra1994@gmail.com").Folders("Inbox").Folders("My emails")  # Adjust with your email and folder name

        # Get all unread items in the inbox at once
        unread_messages = inbox.Items.Restrict("[Unread]=True")  # Restrict to unread emails

        # Ensure all unread emails are processed at once
        if unread_messages.Count > 0:
            process_unread_emails(unread_messages, sheet)
        else:
            messagebox.showinfo("No Unread Emails", "There are no unread emails to process.")

        # Save the updated workbook
        wb.save(extracted_email_file)
        messagebox.showinfo("Success", "All unread emails have been processed and saved!")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to connect to Outlook or process emails: {e}")

# Function to search data based on the name (column "K")
# # Function to search data based on the name (column "K")
def search_data():
    name_to_search = entry_search.get()
    if not name_to_search:
        messagebox.showerror("Input Error", "Please enter a name to search.")
        return

    # Open the extracted email data file (job_openings_data.xlsx)
    wb = openpyxl.load_workbook(extracted_email_file)
    sheet = wb.active

    # Iterate through rows to search for the name in column K (index 10, since 0-based index)
    found = False
    for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, values_only=True):
        if row[2] and row[2].lower() == name_to_search.lower():  # Column K is index 10
            entry_name.delete(0, tk.END)
            Received_from.delete(0, tk.END)
            Action_Required.delete("1.0", tk.END)  # Use "1.0" for Text widget
            Summary.delete("1.0", tk.END)  # Use "1.0" for Text widget

            entry_name.insert(0, row[2])  # Name from column K
            Action_Required.insert("1.0", row[4])  # Company Name (adjust if in a different column)
            Summary.insert("1.0", row[5])  # Skills Required (adjust if in a different column)
            Work_Package.insert(0, row[3])
            found = True
            break

    if not found:
        messagebox.showerror("Not Found", f"No data found for '{name_to_search}'.")

# Function to submit updated data
def submit_data():
    # Fetching all the inputs
    name = entry_name.get()
    ACF2_ID = label_ACF2ID.get()
    team_selected = combo.get()
    Policy = Policy_number.get()
    WP = Work_Package.get()
    trigger = Trigger.get()
    Receivedfrom = Received_from.get() 
    dollar = Dollar_Impact.get()
    recovery = Service_Recovery.get() 
    work_type = Action_Required.get("1.0", tk.END)  # Retrieving full text
    comments = Summary.get("1.0", tk.END)  # Retrieving full text

    # Input validation
    if not name or not Receivedfrom or not work_type or not comments:
        messagebox.showerror("Input Error", "Please fill out all fields.")
        return

    # Open the updated file or create a new one if it doesn't exist
    if os.path.exists(updated_data_file):
        wb = openpyxl.load_workbook(updated_data_file)
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        # Adding headers if a new file is created
        headers = [ "CSR Name","ACF2 ID","Functional Team", "Policy Number", "Work Package", "Trigger Type", 
                   "Received From", "Dollar Impact", "Service Recovery", 'Action Required', 'Summary', ]
        sheet.append(headers)
    sheet = wb.active

    # Appending the new/updated data into respective columns
    sheet.append([name,ACF2_ID,team_selected ,Policy, WP, trigger, Receivedfrom, dollar, recovery, work_type.strip(), comments.strip()])

    # Save the workbook
    wb.save(updated_data_file)
    messagebox.showinfo("Success", "Data submitted successfully!")

# Function to refresh (clear) the entry fields
def refresh_entries():
    entry_name.delete(0, tk.END)
    Received_from.delete(0, tk.END)
    label_ACF2ID.delete(0, tk.END)
    Policy_number.delete(0,tk.END)
    Work_Package.delete(0,tk.END)
    Trigger.delete(0,tk.END)
    Dollar_Impact.delete(0,tk.END)
    Service_Recovery.delete(0,tk.END)

    Action_Required.delete("1.0", tk.END)
    Summary.delete("1.0", tk.END)
    entry_search.delete(0, tk.END)


# Create the GUI window
import tkinter as tk

root = tk.Tk()

root.title("LO Tracker")

# root.title("Data Tool")
root.geometry("900x410")
root.resizable(0,0)
# root.config(bg="#2F4F4F")
root.config(bg="#d2f8d2")


heading = Label(root, text="LO Tracker", font=('Times', 12, ), pady=2, bd=5, width=99, bg="#d2f8d2", fg="#030303",)
heading.grid(row=0, column=0, )

F1 = LabelFrame(root, text="CSR Details", font=('times', 10,), fg="#030303", bg="#d2f8d2")
F1.place(x=0, y=25, relwidth=1)

# Create input fields and labels
label_search = tk.Label(F1, text="Search by Name", bg="#d2f8d2",fg="#030303", font=('times new roman', 13))
label_search.grid(row=0, column=0, padx=20, pady=5)

entry_search = tk.Entry(F1,fg="#030303",bg="White",font=('Times',13), bd=5, relief=GROOVE, width=15, )
entry_search.grid(row=0, column=1, )

label_name = tk.Label(F1, text="CSR Name", bg="#d2f8d2",fg="#030303", font=('Times',13))
label_name.grid(row=0,column=2, padx=22, pady=5)

entry_name = tk.Entry(F1, fg="#030303",bg="White",font=('Times',13), bd=5, relief=GROOVE, width=15,)
entry_name.grid(row=0, column=3)

label_ACF2ID = tk.Label(F1, text=" ACF2 ID", bg="#d2f8d2",fg="#030303", font=('Times',13))
label_ACF2ID.grid(row=0,column=4, padx=38, pady=5)

label_ACF2ID = tk.Entry(F1, fg="#030303",bg="White",font=('Times',13), bd=5, relief=GROOVE, width=15,)
label_ACF2ID.grid(row=0, column=5)


label_Team = tk.Label(F1, text="Functional Team", bg="#d2f8d2",fg="#030303", font=('Times',13))
label_Team.grid(row=1, column=0, padx=20, pady=5)

team = [" ",'Ins Admin-SLGS', 'Payment-SLGS', 'Plan Change-SLGS', 'Policy Title-SLGS', 'Taxation-SLGS']
combo=ttk.Combobox(F1 ,values=team,width=15, height=1,  style="TCombobox" )
combo.grid(row=1,column=1, sticky="w",)


F2 = LabelFrame(root, text="Error Details", font=('times', 10,), fg="#030303", bg="#d2f8d2")
F2.place(x=0, y=112, relwidth=1)

Policy_number = tk.Label(F2, text="Policy Number", bg="#d2f8d2",fg="#030303", font=('times new roman', 13))
Policy_number.grid(row=0,column=0, padx=20, pady=5)

Policy_number = tk.Entry(F2,fg="#030303",bg="White",font=('Times',13), bd=5,relief=GROOVE, width=15, )
Policy_number.grid(row=0, column=1, padx=10)


Work_Package = tk.Label(F2, text="Work Package", bg="#d2f8d2",fg="#030303", font=('times new roman', 13))
Work_Package .grid(row=0,column=2,  pady=5)

Work_Package  = tk.Entry(F2,fg="#030303",bg="White",font=('Times',13), bd=5, relief=GROOVE,  width=15, )
Work_Package .grid(row=0, column=3,padx=10)

Trigger = tk.Label(F2, text="Trigger Type", bg="#d2f8d2",fg="#030303", font=('times new roman', 13))
Trigger .grid(row=0,column=4,  pady=5)

Trigger  = tk.Entry(F2,fg="#030303",bg="White",font=('Times',13), bd=5, relief=GROOVE, width=15, )
Trigger .grid(row=0, column=5,padx=10)


Received_from_label = tk.Label(F2, text="Received From", bg="#d2f8d2",fg="#030303", font=('times new roman', 13))
Received_from_label .grid(row=2,column=0, padx=20, pady=5)

Received_from = tk.Entry(F2,fg="#030303",bg="White",font=('Times',13), bd=5, relief=GROOVE,  width=15, )
Received_from .grid(row=2, column=1, padx=10)


Dollar_Impact = tk.Label(F2, text="Dollar Impact", bg="#d2f8d2",fg="#030303", font=('times new roman', 13))
Dollar_Impact .grid(row=2,column=2,  pady=5)

Dollar_Impact  = tk.Entry(F2,fg="#030303",bg="White",font=('Times',13), bd=5, relief=GROOVE, width=15, )
Dollar_Impact .grid(row=2, column=3,padx=10)


Service_Recovery = tk.Label(F2, text="Service Recovery", bg="#d2f8d2",fg="#030303", font=('times new roman', 13))
Service_Recovery .grid(row=2,column=4, pady=5)

Service_Recovery  = tk.Entry(F2,fg="#030303",bg="White",font=('Times',13), bd=5, relief=GROOVE,width=15 )
Service_Recovery .grid(row=2, column=5,padx=10)

F3 = LabelFrame(root, text="Discription", font=('times', 10,),  fg="#030303", bg="#d2f8d2")
F3.place(x=0, y=199, relwidth=1)

Action_Required_Label = tk.Label(F3, text="Action Required", bg="#d2f8d2",fg="#030303", font=('times new roman', 13),bd=2, width=15, height=1 )
Action_Required_Label.grid(row=0,column=0, padx=0, pady=5)

Action_Required = tk.Text(F3, height=2, width=78,fg="#030303",bg="White", font='Times 13', bd=5,relief=GROOVE )
Action_Required.grid(row=0, column=1, padx=0,pady=0)

Summary_Label = tk.Label(F3, text="Summary", bg="#d2f8d2",fg="#030303",  font=('Times',13) ,bd=2, width=15, height=1)
Summary_Label.grid(row=2,column=0, padx=10, pady=2)

Summary = tk.Text(F3, height=2, width=78,fg="#030303",bg="White",font='Times 13', bd=5,relief=GROOVE )
Summary.grid(row=2, column=1, padx=0, pady=2)


F4 = LabelFrame(root, font=('times', 10,),  fg="#030303", bg="#d2f8d2")
F4.place(x=0, y=321, relwidth=1)

button_get_emails = tk.Button(F4, text="Get Emails",width=12,height=2,bg="#2F4F4F",fg="alice blue",relief= RAISED ,font='arial 10',bd=6,  command=get_emails)
button_get_emails.grid(row=0, column=2,padx=15,pady=15)

button_search = tk.Button(F4, text="Search",width=12,height=2,bg="#2F4F4F",fg="alice blue",relief= RAISED ,font='arial 10',bd=6, command=search_data)
button_search.grid(row=0, column=3, pady=15,padx=15)


button_submit = tk.Button(F4, text="Submit",width=12,height=2,bg="#2F4F4F",fg="alice blue",relief= RAISED ,font='arial 10',bd=6, command=submit_data)
button_submit.grid(row=0, column=4, pady=15,padx=15)


button_refresh = tk.Button(F4, text="Refresh Entries",width=12,height=2,bg="#2F4F4F",fg="alice blue",relief= RAISED ,font='arial 10',bd=6, command=refresh_entries)
button_refresh.grid(row=0, column=5, padx=15, pady=15, )    

blank_label = tk.Label(F4, bg="#d2f8d2",fg="#030303",  font=('Times',13) ,bd=2, width=17, height=1)
blank_label.grid(row=0,column=0,pady=15)

root.mainloop()
